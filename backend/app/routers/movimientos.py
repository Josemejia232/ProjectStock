from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc
from sqlalchemy.orm import selectinload
from typing import List, Optional
import io
import openpyxl

from app.database import get_db
from app.models import Movimiento, Inventario, Proyecto, Material
from app.schemas import MovimientoCreate, MovimientoUpdate, MovimientoOut

router = APIRouter(prefix="/api/movimientos", tags=["Movimientos"])


def build_out(movimiento, proyecto_nombre=None, material_nombre=None, material_unidad=None):
    out = MovimientoOut.model_validate(movimiento)
    out.proyecto_nombre = proyecto_nombre or (movimiento.proyecto.nombre if movimiento.proyecto else None)
    out.material_nombre = material_nombre or (movimiento.material.nombre if movimiento.material else None)
    out.material_unidad = material_unidad or (movimiento.material.unidad_medida if movimiento.material else None)
    return out


async def actualizar_cantidad(db: AsyncSession, proyecto_id: int, material_id: int, tipo: str, cantidad: float):
    result = await db.execute(
        select(Inventario).where(
            Inventario.proyecto_id == proyecto_id,
            Inventario.material_id == material_id,
        )
    )
    inventario = result.scalar_one_or_none()
    if not inventario:
        raise HTTPException(status_code=400, detail="El material no está registrado en el inventario del proyecto")

    if tipo == "entrada":
        inventario.cantidad_actual += cantidad
    elif tipo == "salida":
        if inventario.cantidad_actual < cantidad:
            raise HTTPException(status_code=400, detail="Cantidad insuficiente en inventario para esta salida")
        inventario.cantidad_actual -= cantidad

    await db.flush()
    return inventario


@router.get("/", response_model=List[MovimientoOut])
async def listar_movimientos(
    proyecto_id: Optional[int] = None,
    material_id: Optional[int] = None,
    tipo: Optional[str] = None,
    categoria: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    stmt = (
        select(Movimiento, Proyecto.nombre, Material.nombre, Material.unidad_medida)
        .join(Proyecto, Movimiento.proyecto_id == Proyecto.id)
        .join(Material, Movimiento.material_id == Material.id)
        .order_by(desc(Movimiento.fecha))
    )
    if proyecto_id:
        stmt = stmt.where(Movimiento.proyecto_id == proyecto_id)
    if material_id:
        stmt = stmt.where(Movimiento.material_id == material_id)
    if tipo:
        stmt = stmt.where(Movimiento.tipo == tipo)
    if categoria:
        stmt = stmt.where(Movimiento.categoria == categoria)

    result = await db.execute(stmt)
    rows = result.all()
    return [build_out(mov, proy_nombre, mat_nombre, mat_unidad) for mov, proy_nombre, mat_nombre, mat_unidad in rows]


@router.get("/exportar")
async def exportar_movimientos(
    proyecto_id: Optional[int] = None,
    material_id: Optional[int] = None,
    tipo: Optional[str] = None,
    categoria: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    stmt = (
        select(Movimiento, Proyecto.nombre, Material.nombre, Material.unidad_medida)
        .join(Proyecto, Movimiento.proyecto_id == Proyecto.id)
        .join(Material, Movimiento.material_id == Material.id)
        .order_by(desc(Movimiento.fecha))
    )
    if proyecto_id:
        stmt = stmt.where(Movimiento.proyecto_id == proyecto_id)
    if material_id:
        stmt = stmt.where(Movimiento.material_id == material_id)
    if tipo:
        stmt = stmt.where(Movimiento.tipo == tipo)
    if categoria:
        stmt = stmt.where(Movimiento.categoria == categoria)

    result = await db.execute(stmt)
    rows = result.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    headers = ["Fecha", "No. Remisión", "Insumo", "Cantidad", "Unidad", "Descripción", "Categoría", "Tipo", "Proyecto", "Usuario"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    for i, (mov, proy_nombre, mat_nombre, mat_unidad) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=mov.fecha.strftime("%Y-%m-%d") if mov.fecha else "")
        ws.cell(row=i, column=2, value=mov.no_remision or "")
        ws.cell(row=i, column=3, value=mat_nombre or "")
        ws.cell(row=i, column=4, value=mov.cantidad)
        ws.cell(row=i, column=5, value=mat_unidad or "")
        ws.cell(row=i, column=6, value=mov.descripcion or "")
        ws.cell(row=i, column=7, value=mov.categoria or "")
        ws.cell(row=i, column=8, value=mov.tipo)
        ws.cell(row=i, column=9, value=proy_nombre or "")
        ws.cell(row=i, column=10, value=mov.usuario or "")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=movimientos.xlsx"},
    )


@router.get("/{movimiento_id}", response_model=MovimientoOut)
async def obtener_movimiento(movimiento_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Movimiento, Proyecto.nombre, Material.nombre, Material.unidad_medida)
        .join(Proyecto, Movimiento.proyecto_id == Proyecto.id)
        .join(Material, Movimiento.material_id == Material.id)
        .where(Movimiento.id == movimiento_id)
    )
    row = result.one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    mov, proy_nombre, mat_nombre, mat_unidad = row
    return build_out(mov, proy_nombre, mat_nombre, mat_unidad)


@router.post("/", response_model=MovimientoOut, status_code=status.HTTP_201_CREATED)
async def crear_movimiento(data: MovimientoCreate, db: AsyncSession = Depends(get_db)):
    await actualizar_cantidad(db, data.proyecto_id, data.material_id, data.tipo, data.cantidad)

    movimiento = Movimiento(**data.model_dump())
    db.add(movimiento)
    await db.commit()
    await db.refresh(movimiento)

    result = await db.execute(
        select(Proyecto.nombre, Material.nombre, Material.unidad_medida)
        .where(Proyecto.id == movimiento.proyecto_id, Material.id == movimiento.material_id)
    )
    proy_nombre, mat_nombre, mat_unidad = result.one()
    return build_out(movimiento, proy_nombre, mat_nombre, mat_unidad)


@router.put("/{movimiento_id}", response_model=MovimientoOut)
async def actualizar_movimiento(movimiento_id: int, data: MovimientoUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Movimiento, Proyecto.nombre, Material.nombre, Material.unidad_medida)
        .join(Proyecto, Movimiento.proyecto_id == Proyecto.id)
        .join(Material, Movimiento.material_id == Material.id)
        .where(Movimiento.id == movimiento_id)
    )
    row = result.one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    mov, proy_nombre, mat_nombre, mat_unidad = row
    for key, val in data.model_dump(exclude_unset=True).items():
        setattr(mov, key, val)
    await db.commit()
    await db.refresh(mov)
    return build_out(mov)


@router.delete("/{movimiento_id}", status_code=status.HTTP_204_NO_CONTENT)
async def eliminar_movimiento(movimiento_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Movimiento).options(selectinload(Movimiento.proyecto), selectinload(Movimiento.material))
        .where(Movimiento.id == movimiento_id)
    )
    mov = result.scalar_one_or_none()
    if not mov:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")

    # Revertir inventario
    inv_result = await db.execute(
        select(Inventario).where(
            Inventario.proyecto_id == mov.proyecto_id,
            Inventario.material_id == mov.material_id,
        )
    )
    inventario = inv_result.scalar_one_or_none()
    if inventario:
        if mov.tipo == "entrada":
            inventario.cantidad_actual = max(0, inventario.cantidad_actual - mov.cantidad)
        elif mov.tipo == "salida":
            inventario.cantidad_actual += mov.cantidad

    await db.delete(mov)
    await db.commit()
