from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, desc
from typing import List, Optional, Dict
import io
import openpyxl

from app.database import get_db
from app.models import Inventario, Proyecto, Material, Movimiento, PrecioHistorico
from sqlalchemy import func
from app.schemas import AlertaOut, DashboardOut, AlertaPrecioOut, ResumenInsumoOut, InsumoPorProyectoOut

router = APIRouter(prefix="/api/reportes", tags=["Reportes"])


@router.get("/dashboard", response_model=DashboardOut)
async def get_dashboard(db: AsyncSession = Depends(get_db)):
    proyectos = await db.execute(select(func.count(Proyecto.id)))
    total_proyectos = proyectos.scalar() or 0

    activos = await db.execute(
        select(func.count(Proyecto.id)).where(Proyecto.estado == "ejecucion")
    )
    proyectos_activos = activos.scalar() or 0

    materiales = await db.execute(select(func.count(Material.id)))
    total_materiales = materiales.scalar() or 0

    movs = await db.execute(select(func.count(Movimiento.id)))
    total_movimientos = movs.scalar() or 0

    result = await db.execute(
        select(Inventario, Proyecto.nombre, Material.nombre, Material.unidad_medida)
        .join(Proyecto, Inventario.proyecto_id == Proyecto.id)
        .join(Material, Inventario.material_id == Material.id)
    )
    rows = result.all()

    alertas = []
    excedentes = 0
    faltantes = 0
    stock_critico = 0

    for inv, proy_nombre, mat_nombre, mat_unidad in rows:
        tipo = None
        if inv.cantidad_maxima > 0 and inv.cantidad_actual > inv.cantidad_maxima:
            tipo = "excedente"
            excedentes += 1
        elif inv.cantidad_minima > 0 and inv.cantidad_actual < inv.cantidad_minima:
            tipo = "faltante"
            faltantes += 1
        elif inv.cantidad_actual == 0:
            tipo = "stock_critico"
            stock_critico += 1

        if tipo:
            alertas.append(AlertaOut(
                inventario_id=inv.id,
                proyecto_id=inv.proyecto_id,
                proyecto_nombre=proy_nombre,
                material_id=inv.material_id,
                material_nombre=mat_nombre,
                material_unidad=mat_unidad,
                cantidad_actual=inv.cantidad_actual,
                cantidad_minima=inv.cantidad_minima,
                cantidad_maxima=inv.cantidad_maxima,
                tipo_alerta=tipo,
            ))

    return DashboardOut(
        total_proyectos=total_proyectos,
        proyectos_activos=proyectos_activos,
        total_materiales=total_materiales,
        total_movimientos=total_movimientos,
        alertas=alertas,
        excedentes=excedentes,
        faltantes=faltantes,
        stock_critico=stock_critico,
    )


@router.get("/alertas", response_model=List[AlertaOut])
async def get_alertas(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Inventario, Proyecto.nombre, Material.nombre, Material.unidad_medida)
        .join(Proyecto, Inventario.proyecto_id == Proyecto.id)
        .join(Material, Inventario.material_id == Material.id)
        .where(
            (Inventario.cantidad_actual == 0) |
            ((Inventario.cantidad_minima > 0) & (Inventario.cantidad_actual < Inventario.cantidad_minima)) |
            ((Inventario.cantidad_maxima > 0) & (Inventario.cantidad_actual > Inventario.cantidad_maxima))
        )
    )
    rows = result.all()

    alertas = []
    for inv, proy_nombre, mat_nombre, mat_unidad in rows:
        tipo = "excedente" if (inv.cantidad_maxima > 0 and inv.cantidad_actual > inv.cantidad_maxima) else \
               "faltante" if (inv.cantidad_minima > 0 and inv.cantidad_actual < inv.cantidad_minima) else \
               "stock_critico"
        alertas.append(AlertaOut(
            inventario_id=inv.id,
            proyecto_id=inv.proyecto_id,
            proyecto_nombre=proy_nombre,
            material_id=inv.material_id,
            material_nombre=mat_nombre,
            material_unidad=mat_unidad,
            cantidad_actual=inv.cantidad_actual,
            cantidad_minima=inv.cantidad_minima,
            cantidad_maxima=inv.cantidad_maxima,
            tipo_alerta=tipo,
        ))
    return alertas


@router.get("/alertas-precios", response_model=List[AlertaPrecioOut])
async def get_alertas_precios(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(PrecioHistorico, Material.nombre, Material.proveedor)
        .join(Material, PrecioHistorico.material_id == Material.id)
        .order_by(desc(PrecioHistorico.fecha))
    )
    rows = result.all()

    return [
        AlertaPrecioOut(
            id=h.id,
            material_id=h.material_id,
            material_nombre=mat_nombre,
            proveedor=mat_proveedor or "",
            precio_anterior=h.precio_anterior,
            precio_nuevo=h.precio_nuevo,
            diferencia=round(h.precio_nuevo - h.precio_anterior, 2),
            fecha=h.fecha,
        )
        for h, mat_nombre, mat_proveedor in rows
    ]


@router.get("/resumen-insumos", response_model=List[ResumenInsumoOut])
async def resumen_insumos(proyecto_id: Optional[int] = None, db: AsyncSession = Depends(get_db)):
    mov_filter = []
    inv_filter = []
    if proyecto_id:
        mov_filter.append(Movimiento.proyecto_id == proyecto_id)
        inv_filter.append(Inventario.proyecto_id == proyecto_id)

    entradas_q = select(Movimiento.material_id, func.sum(Movimiento.cantidad).label("total")).where(Movimiento.tipo == "entrada")
    if mov_filter:
        entradas_q = entradas_q.where(*mov_filter)
    entradas_q = entradas_q.group_by(Movimiento.material_id)

    salidas_q = select(Movimiento.material_id, func.sum(Movimiento.cantidad).label("total")).where(Movimiento.tipo == "salida")
    if mov_filter:
        salidas_q = salidas_q.where(*mov_filter)
    salidas_q = salidas_q.group_by(Movimiento.material_id)

    stock_q = select(Inventario.material_id, func.sum(Inventario.cantidad_actual).label("total"))
    if inv_filter:
        stock_q = stock_q.where(*inv_filter)
    stock_q = stock_q.group_by(Inventario.material_id)

    entradas = {r.material_id: r.total for r in (await db.execute(entradas_q)).all()}
    salidas = {r.material_id: r.total for r in (await db.execute(salidas_q)).all()}
    stock = {r.material_id: r.total for r in (await db.execute(stock_q)).all()}

    all_ids = set(entradas) | set(salidas) | set(stock)
    materials_map = {}
    if all_ids:
        mats = await db.execute(select(Material).where(Material.id.in_(all_ids)))
        materials_map = {m.id: m for m in mats.scalars().all()}

    result = []
    for mid in sorted(all_ids):
        mat = materials_map.get(mid)
        result.append(ResumenInsumoOut(
            material_id=mid,
            material_nombre=mat.nombre if mat else f"Material #{mid}",
            unidad_medida=mat.unidad_medida if mat else "",
            entradas=float(entradas.get(mid, 0)),
            salidas=float(salidas.get(mid, 0)),
            stock=float(stock.get(mid, 0)),
        ))
    return result


@router.get("/exportar-insumos")
async def exportar_insumos(proyecto_id: Optional[int] = None, db: AsyncSession = Depends(get_db)):
    mov_filter = []
    inv_filter = []
    if proyecto_id:
        mov_filter.append(Movimiento.proyecto_id == proyecto_id)
        inv_filter.append(Inventario.proyecto_id == proyecto_id)

    entradas_q = select(Movimiento.material_id, func.sum(Movimiento.cantidad).label("total")).where(Movimiento.tipo == "entrada")
    if mov_filter:
        entradas_q = entradas_q.where(*mov_filter)
    entradas_q = entradas_q.group_by(Movimiento.material_id)

    salidas_q = select(Movimiento.material_id, func.sum(Movimiento.cantidad).label("total")).where(Movimiento.tipo == "salida")
    if mov_filter:
        salidas_q = salidas_q.where(*mov_filter)
    salidas_q = salidas_q.group_by(Movimiento.material_id)

    stock_q = select(Inventario.material_id, func.sum(Inventario.cantidad_actual).label("total"))
    if inv_filter:
        stock_q = stock_q.where(*inv_filter)
    stock_q = stock_q.group_by(Inventario.material_id)

    entradas = {r.material_id: r.total for r in (await db.execute(entradas_q)).all()}
    salidas = {r.material_id: r.total for r in (await db.execute(salidas_q)).all()}
    stock = {r.material_id: r.total for r in (await db.execute(stock_q)).all()}

    all_ids = set(entradas) | set(salidas) | set(stock)
    materials_map = {}
    if all_ids:
        mats = await db.execute(select(Material).where(Material.id.in_(all_ids)))
        materials_map = {m.id: m for m in mats.scalars().all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Insumos"
    headers = ["Insumo", "Unidad", "Entradas", "Salidas", "Stock"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    row_num = 2
    for mid in sorted(all_ids):
        mat = materials_map.get(mid)
        ws.cell(row=row_num, column=1, value=mat.nombre if mat else f"Material #{mid}")
        ws.cell(row=row_num, column=2, value=mat.unidad_medida if mat else "")
        ws.cell(row=row_num, column=3, value=float(entradas.get(mid, 0)))
        ws.cell(row=row_num, column=4, value=float(salidas.get(mid, 0)))
        ws.cell(row=row_num, column=5, value=float(stock.get(mid, 0)))
        row_num += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=resumen_insumos.xlsx"},
    )


@router.get("/movimientos-por-mes")
async def movimientos_por_mes(db: AsyncSession = Depends(get_db)):
    from sqlalchemy import extract
    result = await db.execute(
        select(
            extract("year", Movimiento.fecha).label("year"),
            extract("month", Movimiento.fecha).label("month"),
            Movimiento.tipo,
            func.count(Movimiento.id).label("total"),
        )
        .group_by(
            extract("year", Movimiento.fecha),
            extract("month", Movimiento.fecha),
            Movimiento.tipo,
        )
        .order_by(desc(extract("year", Movimiento.fecha)), desc(extract("month", Movimiento.fecha)))
    )
    rows = result.all()
    data = {}
    for year_val, month_val, tipo, total in rows:
        mes = f"{int(year_val)}-{int(month_val):02d}"
        if mes not in data:
            data[mes] = {"mes": mes, "entradas": 0, "salidas": 0}
        data[mes][tipo + "s"] = total
    return list(data.values())


@router.get("/insumos-por-proyecto", response_model=InsumoPorProyectoOut)
async def insumos_por_proyecto(db: AsyncSession = Depends(get_db)):
    proy_result = await db.execute(select(Proyecto).order_by(Proyecto.nombre))
    proyectos = proy_result.scalars().all()

    result = await db.execute(
        select(Inventario, Material.nombre, Material.unidad_medida)
        .join(Material, Inventario.material_id == Material.id)
    )
    rows = result.all()

    # material_id -> { nombre, unidad, {proyecto_id: cantidad} }
    data: Dict[int, dict] = {}
    for inv, mat_nombre, mat_unidad in rows:
        mid = inv.material_id
        if mid not in data:
            data[mid] = {"nombre": mat_nombre, "unidad": mat_unidad, "por_proyecto": {}}
        data[mid]["por_proyecto"][inv.proyecto_id] = inv.cantidad_actual

    proy_nombres = [p.nombre for p in proyectos]
    insumos = []
    for mid in sorted(data.keys()):
        d = data[mid]
        cantidades: Dict[str, float] = {}
        total = 0.0
        for p in proyectos:
            c = d["por_proyecto"].get(p.id, 0.0)
            cantidades[p.nombre] = c
            total += c
        insumos.append({
            "material_id": mid,
            "material_nombre": d["nombre"],
            "unidad_medida": d["unidad"],
            "cantidades": cantidades,
            "total": total,
        })

    return {"proyectos": proy_nombres, "insumos": insumos}
