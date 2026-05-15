from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import List
import io
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone

from app.database import get_db
from app.models import Material, PrecioHistorico
from app.schemas import MaterialCreate, MaterialUpdate, MaterialOut

router = APIRouter(prefix="/api/materiales", tags=["Materiales"])


@router.get("/", response_model=List[MaterialOut])
async def listar_materiales(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Material).order_by(Material.nombre))
    return result.scalars().all()


@router.get("/plantilla")
async def descargar_plantilla():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Insumos"
    headers = ["Nombre", "Unidad", "Categoria", "Precio Unitario", "Proveedor"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 25

    data = [["Cemento Portland", "kg", "Concreto", 12.50, "Cementos Ejemplo"]]
    for row in data:
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_insumos.xlsx"},
    )


@router.post("/cargar")
async def cargar_desde_excel(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Get existing material names
    existing = await db.execute(select(Material.nombre))
    existing_names = {r.lower().strip() for r in existing.scalars().all()}

    creados = 0
    omitidos = 0
    for row in rows:
        nombre = str(row[0]).strip() if row[0] else ""
        if not nombre:
            continue
        if nombre.lower().strip() in existing_names:
            omitidos += 1
            continue

        unidad = str(row[1]).strip() if row[1] else ""
        categoria = str(row[2]).strip() if row[2] else "General"
        precio = float(row[3]) if row[3] else 0.0
        proveedor = str(row[4]).strip() if len(row) > 4 and row[4] else ""

        material = Material(
            nombre=nombre,
            unidad_medida=unidad or "ud",
            categoria=categoria,
            precio_unitario=precio,
            proveedor=proveedor,
        )
        db.add(material)
        creados += 1

    await db.commit()
    return {"creados": creados, "omitidos": omitidos}


@router.get("/{material_id}", response_model=MaterialOut)
async def obtener_material(material_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Material).where(Material.id == material_id))
    material = result.scalar_one_or_none()
    if not material:
        raise HTTPException(status_code=404, detail="Material no encontrado")
    return material


@router.post("/", response_model=MaterialOut, status_code=status.HTTP_201_CREATED)
async def crear_material(data: MaterialCreate, db: AsyncSession = Depends(get_db)):
    material = Material(**data.model_dump())
    db.add(material)
    try:
        await db.commit()
        await db.refresh(material)
        return material
    except Exception:
        await db.rollback()
        raise HTTPException(status_code=400, detail="El material ya existe o datos inválidos")


@router.put("/{material_id}", response_model=MaterialOut)
async def actualizar_material(material_id: int, data: MaterialUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Material).where(Material.id == material_id))
    material = result.scalar_one_or_none()
    if not material:
        raise HTTPException(status_code=404, detail="Material no encontrado")

    old_price = material.precio_unitario
    for key, val in data.model_dump(exclude_unset=True).items():
        setattr(material, key, val)

    if data.precio_unitario is not None and old_price != data.precio_unitario:
        historico = PrecioHistorico(
            material_id=material_id,
            precio_anterior=old_price,
            precio_nuevo=data.precio_unitario,
        )
        db.add(historico)

    await db.commit()
    await db.refresh(material)
    return material


@router.delete("/{material_id}", status_code=status.HTTP_204_NO_CONTENT)
async def eliminar_material(material_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Material).where(Material.id == material_id))
    material = result.scalar_one_or_none()
    if not material:
        raise HTTPException(status_code=404, detail="Material no encontrado")
    await db.delete(material)
    await db.commit()
